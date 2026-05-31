#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
'시각용 조편성 엑셀' (계층형, 첨부 '대학 3일차 조편성.xlsx' 형태)
  → 'PaceRise 시스템 업로드용 평탄형 조편성' 엑셀

읽는 양식 (입력):
  A1: 카테고리 (A:I 병합, 예 '남자 대학부' / '여자 대학부')
  A2: ▣ 종목명 (결승)        — 종목 헤더 (앞에 ▣)
  A3: '1조' / '2조' …         — 조 라벨
  B3..E3: 표 헤더 ('순'/'레인' | '번호' | '성명' | '소속')
  B4..E4부터 데이터:
      - 일반: B=순/레인, C=번호, D=성명, E=소속
      - 릴레이: 같은 팀이면 B(레인)+E(소속)는 첫 줄에만, 나머지는 C,D만

쓰는 양식 (출력):
  성별 | 종목 | 라운드 | 조 | 그룹 | 순서 | 배번 | 성명 | 소속
  (server.js parseHeatAssignmentExcel 가 그대로 읽는 평탄형)

특이 처리:
- '▣ 여자대학교부' 같이 ▣ 로 시작하지만 종목명이 아닌 카테고리 보조 라벨은 스킵
- 데일리에 있던 'A,B 조' (▣ A / ▣ B) 표기는 대학 파일에선 안 쓰임 — 대신
  여러 결승 그룹은 그냥 '1조', '2조' 로 분리되어 있음 → heat 가 1,2,3…
  으로 증가
  (단, 장거리(5000m, 3000m, 1500m, 10000m, 3000mSC) + 조가 2개 이상이면
   실업부 데일리와 동일하게 'A/B 그룹 결승'으로 간주 → heat=1, group=A/B 로 출력)
"""

import sys, re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# 평탄형 출력 헬퍼는 데일리 변환기와 공유
sys.path.insert(0, '/home/user/webapp/tools')
from convert_daily_to_upload import (
    HEADERS, write_flat_sheet,
    detect_gender, detect_round, clean_event_name,
)
from convert_daily_to_schedule import is_relay, is_track_lane


CATEGORY_RE   = re.compile(r'(남자|여자)?\s*(대학부|실업부|중등부|고등부|일반부|마스터즈|마스터|꿈나무|선수권|U18|U20)')
EVENT_HEADER  = re.compile(r'^▣\s*(.+)$')
GROUP_LABEL   = re.compile(r'^(\d+)\s*조$')
SECTION_LABEL = ('순', '레인')


def is_category(s):
    """A열 셀값이 카테고리(예 '남자 대학부') 인지 판정."""
    if not s:
        return False
    s = str(s).strip()
    # '▣' 로 시작하면 종목 헤더 또는 카테고리 보조 라벨
    if s.startswith('▣'):
        return False
    return bool(CATEGORY_RE.search(s))


def parse_schedule_sheet(ws):
    """시각용 조편성 시트 → events 구조 (convert_daily_to_upload 의 입력과 동일).

    return: [
        {
            'category': '남자 대학부',
            'event':    '5000m (결승)',
            'is_relay': bool,
            'use_lane': bool,
            'groups': [
                {
                    'group_label': '1조',
                    'rows': [
                        # 일반:   {'no_or_lane','bib','name','team'}
                        # 릴레이: {'lane','team','members':[{'bib','name'}]}
                    ]
                }, ...
            ]
        }, ...
    ]
    """
    events = []
    current_category = None
    current_event = None
    current_group = None
    current_relay_team = None  # 릴레이 진행 중인 팀

    nrows = ws.max_row
    for r in range(1, nrows + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        e = ws.cell(row=r, column=5).value

        a_str = str(a).strip() if a is not None else ''
        b_str = str(b).strip() if b is not None else ''

        # --- 1) 카테고리 헤더 (A:I 병합되어 있을 수 있음) ---
        if is_category(a_str):
            current_category = a_str
            current_event = None
            current_group = None
            current_relay_team = None
            continue

        # --- 2) 종목 헤더 (A열에 '▣ ...') ---
        if a_str.startswith('▣'):
            ev_name = a_str.lstrip('▣').strip()
            # 카테고리 보조 라벨 '▣ 여자대학교부' 등 — 종목이 아닌 건 무시
            if not ev_name or CATEGORY_RE.search(ev_name) or '대학교부' in ev_name or '실업부' in ev_name:
                continue
            current_event = {
                'category': current_category,
                'event':    ev_name,
                'is_relay': is_relay(ev_name),
                'use_lane': is_track_lane(ev_name),
                'groups':   []
            }
            events.append(current_event)
            current_group = None
            current_relay_team = None
            continue

        # --- 3) 조 라벨 (A열에 'N조'), 이 행의 B/C/D/E 는 보통 표 헤더(순|번호|성명|소속) ---
        m = GROUP_LABEL.match(a_str)
        if m and current_event is not None:
            current_group = {
                'group_label': a_str,
                'rows': []
            }
            current_event['groups'].append(current_group)
            current_relay_team = None
            # 이 줄은 헤더 행이므로 데이터로 처리하지 않음 (B='순' 또는 '레인')
            continue

        # --- 4) 표 헤더 단독 ('순'/'레인' | '번호' | '성명' | '소속') — 어떤 파일은 조 라벨이 없을 수 있음
        if b_str in SECTION_LABEL and current_event is not None:
            if current_group is None:
                # 조 라벨이 빠진 경우 1조 자동 생성
                current_group = {'group_label': '1조', 'rows': []}
                current_event['groups'].append(current_group)
            current_relay_team = None
            continue

        # --- 5) 데이터 행 ---
        if current_event is None:
            continue
        # 조가 아직 없으면 1조 자동 생성
        if current_group is None:
            current_group = {'group_label': '1조', 'rows': []}
            current_event['groups'].append(current_group)

        # 모든 셀이 비었으면 건너뜀
        if (b is None and c is None and d is None and e is None):
            continue

        if current_event['is_relay']:
            # 릴레이: B(레인) 가 있으면 새 팀, 없으면 같은 팀의 추가 멤버
            if b is not None and str(b).strip() != '':
                current_relay_team = {
                    'lane': str(b).strip(),
                    'team': (str(e).strip() if e is not None else ''),
                    'members': []
                }
                current_group['rows'].append(current_relay_team)
                # 같은 줄에 멤버(C/D)도 있으면 추가
                if c is not None or d is not None:
                    current_relay_team['members'].append({
                        'bib':  '' if c is None else str(c).strip(),
                        'name': '' if d is None else str(d).strip(),
                    })
            else:
                if current_relay_team is None:
                    # 안전장치: 레인 없이 멤버부터 등장하면 무시
                    continue
                if c is None and d is None:
                    continue
                current_relay_team['members'].append({
                    'bib':  '' if c is None else str(c).strip(),
                    'name': '' if d is None else str(d).strip(),
                })
                # 소속이 같은 줄에 있다면 팀 소속이 비었을 때 채워줌
                if not current_relay_team['team'] and e is not None:
                    current_relay_team['team'] = str(e).strip()
        else:
            # 일반 종목
            current_group['rows'].append({
                'no_or_lane': '' if b is None else str(b).strip(),
                'bib':        '' if c is None else str(c).strip(),
                'name':       '' if d is None else str(d).strip(),
                'team':       '' if e is None else str(e).strip(),
            })

    # 데이터 없는 종목은 제거 (예: '여자 대학부 5000m' 가 3명만이라도 있으면 남김)
    events = [ev for ev in events if any(len(g['rows']) > 0 for g in ev['groups'])]
    return events


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else '/tmp/source_format.xlsx'
    out = sys.argv[2] if len(sys.argv) > 2 else '/tmp/out_upload_univ.xlsx'
    sheet = sys.argv[3] if len(sys.argv) > 3 else None

    wb_src = openpyxl.load_workbook(src, data_only=True)
    sh = sheet if (sheet and sheet in wb_src.sheetnames) else wb_src.sheetnames[0]
    events = parse_schedule_sheet(wb_src[sh])

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = '조편성'
    last = write_flat_sheet(ws_out, events)
    wb_out.save(out)
    print(f"[OK] {out}  (행 수: {last - 1})")

    # 요약
    print(f"     종목 수: {len(events)}")
    for ev in events:
        total = sum(
            sum(len(t['members']) for t in g['rows']) if ev['is_relay'] else len(g['rows'])
            for g in ev['groups']
        )
        kind = '릴레이' if ev['is_relay'] else ('레인' if ev['use_lane'] else '순')
        print(f"     - [{ev['category']}] {ev['event']}  | {kind} | 조 {len(ev['groups'])}개 | {total}명")


if __name__ == '__main__':
    main()
