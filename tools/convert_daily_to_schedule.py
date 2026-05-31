#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
'데일리 명단' 엑셀 → 'PaceRise 표준 조편성' 엑셀 변환기

원본(데일리):
  A열  카테고리 (남자실업부 / 여자실업부 / 실업부)
  B열  ▣ 종목명 (결승 ...) / ▣ A / ▣ B
  C열  순/레인
  D열  번호
  E열  성명
  F열  소속

목적(시스템 표준 조편성):
  A열  카테고리 (A:I 병합) / ▣ 종목명 / 1조 / 2조 ...
  B열  순 또는 레인
  C열  번호
  D열  성명
  E열  소속

특이 처리:
  - A/B 조 분리: 데일리에서 '▣ A', '▣ B' 로 표시된 그룹을 각각 '1조', '2조'로
  - 트랙(허들/계주) → '레인', 장거리/필드 → '순'
  - 릴레이(4x400mR / 4600mR Mixed): 한 팀 = 여러 줄 (레인은 첫 줄에만, 번호/이름/소속은 매 줄)
"""

import sys, re, os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy

# ---------- 스타일 정의 (포맷 파일과 동일) ----------
FONT_BASE  = Font(name='맑은 고딕', size=14, bold=True, color='FF000000')
FONT_TITLE = Font(name='맑은 고딕', size=16, bold=True, color='FF000000')
ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ROW_HEIGHT_TITLE = 25.5
ROW_HEIGHT       = 21.0
COL_WIDTHS = {'A': 12, 'B': 8, 'C': 8, 'D': 14, 'E': 26}

def apply_cell(c, value, *, align=ALIGN_LEFT, font=FONT_BASE):
    c.value = value
    c.font = font
    c.alignment = align

# ---------- 종목 분류 ----------
TRACK_HURDLE_RE = re.compile(r'(\d+m?H|hurdle)', re.I)
RELAY_RE        = re.compile(r'(?:4\s*[xX×]\s*\d+|4600m?R|R\s*\(Mixed\))', re.I)
TRACK_SHORT_RE  = re.compile(r'^\s*\d+m\b')  # 100m, 200m, 400m 등 단거리

def is_relay(event):
    return bool(RELAY_RE.search(event)) or 'mR' in event or '계주' in event

def is_track_lane(event):
    # 레인 배정이 필요한 종목: 허들/계주/단거리
    if is_relay(event):
        return True
    if TRACK_HURDLE_RE.search(event):
        return True
    # 단거리 (200m 이하)
    m = re.match(r'^\s*(\d+)\s*m', event)
    if m:
        try:
            dist = int(m.group(1))
            # 800m 이상은 통상 '순'으로 출발
            if dist <= 400:
                return True
        except ValueError:
            pass
    return False

def head_label(event):
    return '레인' if is_track_lane(event) else '순'

def split_ab_label(raw_event):
    """'▣ 5000m (결승 A, B)' → '5000m (결승)' 로 정리. A,B 조 분리는 호출자가 처리."""
    s = raw_event.lstrip('▣').strip()
    # '(결승 A, B)' → '(결승)' / '(예선 A, B, C)' → '(예선)'
    s = re.sub(r'\(\s*(결승|예선|준결승)\s+[A-Z](?:\s*,\s*[A-Z])*\s*\)', r'(\1)', s)
    return s

def normalize_event_title(raw_event):
    """순수 종목 타이틀 (예: '5000m (결승)' 또는 '110mH (결승)')"""
    return split_ab_label(raw_event)


# ---------- 데일리 파서 ----------
def parse_daily_sheet(ws):
    """
    데일리 시트를 (카테고리, 종목, 조 리스트) 구조의 리스트로 파싱.
    return:
        [
          {
            'category': '남자실업부',
            'event'   : '5000m (결승)',
            'is_relay': False,
            'use_lane': False,
            'groups'  : [   # 조 (1조, 2조 …)
                {
                    'group_label': '1조',  # auto-assigned
                    'rows': [
                        # 일반 종목: {'no_or_lane': '1', 'bib':'199', 'name':'김태훈', 'team':'국군체육부대'}
                        # 릴레이   : {'lane':'4', 'team':'과천시청', 'members':[{'bib':'33','name':'김승호'}, ...]}
                    ]
                }, ...
            ]
          }, ...
        ]
    """
    events = []
    current_category = None
    current_event_block = None  # 현재 종목 블록 dict
    current_group = None        # 현재 조 dict (일반 종목 한정)
    current_relay_team = None   # 현재 릴레이 팀 dict

    nrows = ws.max_row
    r = 1
    while r <= nrows:
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        e = ws.cell(row=r, column=5).value
        f = ws.cell(row=r, column=6).value

        # 카테고리 (A열)
        if a and isinstance(a, str) and ('실업부' in a or '대학부' in a):
            current_category = a.strip()
            r += 1
            continue

        # 종목 헤더 (B열에 ▣)
        if b and isinstance(b, str) and b.strip().startswith('▣'):
            stripped = b.strip().lstrip('▣').strip()

            # Case 1: '▣ A' / '▣ B' (조 분리)
            m_grp = re.match(r'^([A-Z])\s*$', stripped)
            if m_grp and current_event_block is not None:
                # 새 조 시작
                group_letter = m_grp.group(1)
                group_label = f"{ord(group_letter) - ord('A') + 1}조"
                current_group = {'group_label': group_label, 'rows': []}
                current_event_block['groups'].append(current_group)
                current_relay_team = None
                r += 1
                continue

            # Case 2: 신규 종목 헤더
            event_title = normalize_event_title(b)
            current_event_block = {
                'category': current_category,
                'event': event_title,
                'is_relay': is_relay(event_title),
                'use_lane': is_track_lane(event_title),
                'groups': []
            }
            events.append(current_event_block)
            # 기본 1조 자동 생성 — 다음 '▣ A' 가 나오면 이 1조를 덮어쓰지 않고,
            #   대신 직전 행의 '▣ A' 자체가 새 조를 만들도록 위에서 분기.
            # 다만 단일 결승은 헤더 다음에 바로 헤더행(순/번호/...)이 오므로
            #   일단 1조를 만들지 말고 첫 데이터 행 직전 또는 첫 '▣ A' 진입 시 만든다.
            current_group = None
            current_relay_team = None
            r += 1
            continue

        # 표 헤더 (C='순' 또는 C='레인')
        if c and isinstance(c, str) and c.strip() in ('순', '레인'):
            # 이 줄은 그냥 표 헤더이므로 건너뛴다.
            # 만약 1조가 아직 없다면 (= A/B 분리가 없는 단일 결승) 1조 자동 생성
            if current_event_block is not None and not current_event_block['groups']:
                current_group = {'group_label': '1조', 'rows': []}
                current_event_block['groups'].append(current_group)
                current_relay_team = None
            r += 1
            continue

        # 데이터 행
        if current_event_block is None:
            r += 1
            continue

        # 조가 아직 없다면 1조 자동 생성
        if not current_event_block['groups']:
            current_group = {'group_label': '1조', 'rows': []}
            current_event_block['groups'].append(current_group)
            current_relay_team = None
        if current_group is None:
            current_group = current_event_block['groups'][-1]

        is_rly = current_event_block['is_relay']

        if is_rly:
            # 릴레이: C(레인)+D(번호)+E(이름)+F(팀) 또는 D(번호)+E(이름)만
            if c is not None and str(c).strip() != '':
                # 새 팀 시작
                current_relay_team = {
                    'lane': str(c).strip(),
                    'team': (str(f).strip() if f else ''),
                    'members': []
                }
                current_group['rows'].append(current_relay_team)
                if d is not None or e is not None:
                    current_relay_team['members'].append({
                        'bib': '' if d is None else str(d).strip(),
                        'name': '' if e is None else str(e).strip()
                    })
            else:
                # 같은 팀의 추가 멤버
                if current_relay_team is None:
                    r += 1
                    continue
                if d is None and e is None:
                    r += 1
                    continue
                current_relay_team['members'].append({
                    'bib': '' if d is None else str(d).strip(),
                    'name': '' if e is None else str(e).strip()
                })
        else:
            # 일반 종목
            if c is None and d is None and e is None:
                r += 1
                continue
            current_group['rows'].append({
                'no_or_lane': '' if c is None else str(c).strip(),
                'bib':        '' if d is None else str(d).strip(),
                'name':       '' if e is None else str(e).strip(),
                'team':       '' if f is None else str(f).strip()
            })

        r += 1

    return events


# ---------- 시스템 표준 조편성 시트 작성 ----------
def write_schedule_sheet(ws, events, day_title):
    """events 리스트를 ws에 쓴다."""
    # 컬럼 너비
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # 카테고리별로 그룹핑 (남자실업부 → 여자실업부 → 실업부 순서 유지)
    # 데일리는 이미 그 순서대로 들어옴 → events 순서를 카테고리 우선으로 재정렬
    # 다만 같은 카테고리 내 종목 순서는 보존
    order_seen = []
    cat_to_events = {}
    for ev in events:
        cat = ev['category'] or ''
        if cat not in cat_to_events:
            cat_to_events[cat] = []
            order_seen.append(cat)
        cat_to_events[cat].append(ev)

    row = 1
    for cat in order_seen:
        # === 카테고리 헤더 (A:I 병합) ===
        apply_cell(ws.cell(row=row, column=1), cat,
                   align=ALIGN_CENTER, font=FONT_TITLE)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.row_dimensions[row].height = ROW_HEIGHT_TITLE
        row += 1

        for ev in cat_to_events[cat]:
            # === 종목명 (A열) ===
            apply_cell(ws.cell(row=row, column=1), f"▣ {ev['event']}")
            ws.row_dimensions[row].height = ROW_HEIGHT
            row += 1

            lane_or_seq = '레인' if ev['use_lane'] else '순'

            for gi, grp in enumerate(ev['groups']):
                # === 조 헤더 행 (A열: '1조', B:'순/레인', C:'번호', D:'성명', E:'소속') ===
                apply_cell(ws.cell(row=row, column=1), grp['group_label'])
                apply_cell(ws.cell(row=row, column=2), lane_or_seq)
                apply_cell(ws.cell(row=row, column=3), '번호')
                apply_cell(ws.cell(row=row, column=4), '성명')
                apply_cell(ws.cell(row=row, column=5), '소속')
                ws.row_dimensions[row].height = ROW_HEIGHT
                row += 1

                # === 데이터 행 ===
                if ev['is_relay']:
                    for team in grp['rows']:
                        first_member = True
                        for mem in team['members']:
                            if first_member:
                                apply_cell(ws.cell(row=row, column=2), team['lane'])
                                apply_cell(ws.cell(row=row, column=5), team['team'])
                                first_member = False
                            apply_cell(ws.cell(row=row, column=3), mem['bib'])
                            apply_cell(ws.cell(row=row, column=4), mem['name'])
                            ws.row_dimensions[row].height = ROW_HEIGHT
                            row += 1
                else:
                    for rec in grp['rows']:
                        apply_cell(ws.cell(row=row, column=2), rec['no_or_lane'])
                        apply_cell(ws.cell(row=row, column=3), rec['bib'])
                        apply_cell(ws.cell(row=row, column=4), rec['name'])
                        apply_cell(ws.cell(row=row, column=5), rec['team'])
                        ws.row_dimensions[row].height = ROW_HEIGHT
                        row += 1
        # 카테고리 사이 빈 줄 X (포맷 파일도 빈 줄 없음)

    return row


# ---------- main ----------
def main():
    src = sys.argv[1] if len(sys.argv) > 1 else '/tmp/source_daily.xlsx'
    out = sys.argv[2] if len(sys.argv) > 2 else '/tmp/out_schedule.xlsx'
    day = sys.argv[3] if len(sys.argv) > 3 else '3일차 데일리'

    wb_src = openpyxl.load_workbook(src, data_only=True)
    if day not in wb_src.sheetnames:
        print(f"[ERR] 시트 '{day}' 가 없습니다. 사용 가능: {wb_src.sheetnames}")
        sys.exit(1)
    events = parse_daily_sheet(wb_src[day])

    # === 출력 워크북 ===
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = day.replace('데일리', '조편성').strip() or 'Sheet1'

    write_schedule_sheet(ws_out, events, day_title=day)

    wb_out.save(out)
    print(f"[OK] {out}")
    print(f"     종목 수: {len(events)}")
    for ev in events:
        rels = '릴레이' if ev['is_relay'] else ('레인' if ev['use_lane'] else '순')
        total = sum(
            sum(len(t['members']) for t in g['rows']) if ev['is_relay'] else len(g['rows'])
            for g in ev['groups']
        )
        print(f"     - [{ev['category']}] {ev['event']}  | {rels} | 조 {len(ev['groups'])}개 | 인원/멤버 {total}명")

if __name__ == '__main__':
    main()
