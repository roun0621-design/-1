#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
'데일리 명단' 엑셀 → 'PaceRise 시스템 업로드용 평탄형 조편성' 엑셀 변환기

출력 양식 (server.js parseHeatAssignmentExcel 가 기대하는 형태):
    성별 | 종목 | 라운드 | 조 | 그룹 | 순서 | 배번 | 성명 | 소속
    한 줄에 한 명. 첫 행이 헤더.

A,B조 처리:
  - 한국 육상 관행상 '결승 A, B' 는 1조/2조가 아니라 **한 결승 안에서 그룹이 둘** (조 1, 그룹 A/B).
  - 시스템도 'group' 컬럼을 별도로 받으므로 다음과 같이 매핑:
       데일리 '▣ A' → heat=1, group='A'
       데일리 '▣ B' → heat=1, group='B'
  - 단일 결승은 heat=1, group=None.

릴레이:
  - 한 팀(레인)마다 멤버 여러 명. 시스템은 멤버를 모두 별도 행으로 받으면
    같은 lane 으로 묶인다 (heat=1, lane=4 인 행이 모두 한 팀).
  - 4600mR(Mixed) 도 같은 방식 (성별=X / 혼성).
"""

import sys, re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ---------- 데일리 파서 (조편성 변환기와 공유) ----------
sys.path.insert(0, '/home/user/webapp/tools')
from convert_daily_to_schedule import parse_daily_sheet, is_relay, is_track_lane


def detect_gender(category):
    """카테고리 이름에서 성별 추론."""
    if not category:
        return ''
    if '남자' in category:
        return '남'
    if '여자' in category:
        return '여'
    # '실업부' 처럼 성별 미상이면 혼성 (Mixed)
    return '혼성'


def detect_round(event_title):
    """종목명 끝의 (결승) / (예선) / (준결승) 에서 라운드 추출."""
    m = re.search(r'\((결승|예선|준결승)\)', event_title)
    if m:
        return m.group(1)
    return '결승'


def clean_event_name(event_title):
    """'5000m (결승)' → '5000m' 으로 라운드 표시 제거."""
    return re.sub(r'\s*\([^)]*\)\s*$', '', event_title).strip()


# ---------- 출력 ----------
HEADERS = ['성별', '종목', '라운드', '조', '그룹', '순서', '배번', '성명', '소속']

def write_flat_sheet(ws, events):
    # 헤더
    header_font = Font(name='맑은 고딕', size=12, bold=True, color='FFFFFFFF')
    header_fill = PatternFill('solid', fgColor='FF305496')
    body_font   = Font(name='맑은 고딕', size=11)
    align_c     = Alignment(horizontal='center', vertical='center')
    align_l     = Alignment(horizontal='left',   vertical='center')

    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align_c

    ws.column_dimensions['A'].width = 8   # 성별
    ws.column_dimensions['B'].width = 22  # 종목
    ws.column_dimensions['C'].width = 8   # 라운드
    ws.column_dimensions['D'].width = 6   # 조
    ws.column_dimensions['E'].width = 8   # 그룹
    ws.column_dimensions['F'].width = 8   # 순서/레인
    ws.column_dimensions['G'].width = 10  # 배번
    ws.column_dimensions['H'].width = 14  # 성명
    ws.column_dimensions['I'].width = 26  # 소속
    ws.freeze_panes = 'A2'

    row = 2
    for ev in events:
        gender = detect_gender(ev['category'])
        event_name = clean_event_name(ev['event'])
        round_label = detect_round(ev['event'])
        rly = ev['is_relay']

        for gi, grp in enumerate(ev['groups']):
            # 데일리의 '1조'(A)와 '2조'(B) 처리:
            #   - 데일리에서 A/B 로 표시된 그룹은 '한 결승 안의 두 그룹'으로 봄 → heat=1
            #     group = 'A' or 'B'
            #   - 그게 아니면 단순히 1조, 2조, 3조 ... 로 heat 증가
            label = grp['group_label']  # '1조', '2조', ...
            # A/B 그룹 케이스 판정: 종목 헤더에 '(결승 A, B)' 같은 표기가 있었는지
            # parse 단계에선 이 정보가 normalize 되어 사라졌으므로,
            # 조가 2개 이상이고 종목이 장거리(5000m, 3000m, 1500m, 3000mSC 등)인 경우
            # A/B 그룹으로 간주
            use_ab_group = (len(ev['groups']) >= 2 and
                            re.search(r'(5000m|3000m|1500m|10000m|3000mSC|10,000m)', event_name))

            if use_ab_group:
                heat = 1
                group_letter = chr(ord('A') + gi)
            else:
                heat = gi + 1
                group_letter = ''

            if rly:
                # 릴레이: 한 팀(레인)당 여러 멤버
                for team in grp['rows']:
                    lane = team['lane']
                    team_name = team['team']
                    for mi, mem in enumerate(team['members'], start=1):
                        ws.cell(row=row, column=1, value=gender).alignment = align_c
                        ws.cell(row=row, column=2, value=event_name).alignment = align_l
                        ws.cell(row=row, column=3, value=round_label).alignment = align_c
                        ws.cell(row=row, column=4, value=heat).alignment = align_c
                        ws.cell(row=row, column=5, value=group_letter).alignment = align_c
                        ws.cell(row=row, column=6, value=lane).alignment = align_c
                        ws.cell(row=row, column=7, value=mem['bib']).alignment = align_c
                        ws.cell(row=row, column=8, value=mem['name']).alignment = align_l
                        ws.cell(row=row, column=9, value=team_name).alignment = align_l
                        for col in range(1, 10):
                            ws.cell(row=row, column=col).font = body_font
                        row += 1
            else:
                for rec in grp['rows']:
                    ws.cell(row=row, column=1, value=gender).alignment = align_c
                    ws.cell(row=row, column=2, value=event_name).alignment = align_l
                    ws.cell(row=row, column=3, value=round_label).alignment = align_c
                    ws.cell(row=row, column=4, value=heat).alignment = align_c
                    ws.cell(row=row, column=5, value=group_letter).alignment = align_c
                    ws.cell(row=row, column=6, value=rec['no_or_lane']).alignment = align_c
                    ws.cell(row=row, column=7, value=rec['bib']).alignment = align_c
                    ws.cell(row=row, column=8, value=rec['name']).alignment = align_l
                    ws.cell(row=row, column=9, value=rec['team']).alignment = align_l
                    for col in range(1, 10):
                        ws.cell(row=row, column=col).font = body_font
                    row += 1

    return row - 1  # 마지막 데이터 행 번호


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else '/tmp/source_daily.xlsx'
    out = sys.argv[2] if len(sys.argv) > 2 else '/tmp/out_upload.xlsx'
    day = sys.argv[3] if len(sys.argv) > 3 else '3일차 데일리'

    wb_src = openpyxl.load_workbook(src, data_only=True)
    if day not in wb_src.sheetnames:
        print(f"[ERR] 시트 '{day}' 없음. 가능: {wb_src.sheetnames}")
        sys.exit(1)
    events = parse_daily_sheet(wb_src[day])

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = '조편성'

    last = write_flat_sheet(ws_out, events)
    wb_out.save(out)
    print(f"[OK] {out}  (행 수: {last - 1})")

    # 요약
    total = 0
    for ev in events:
        for grp in ev['groups']:
            if ev['is_relay']:
                total += sum(len(t['members']) for t in grp['rows'])
            else:
                total += len(grp['rows'])
    print(f"     총 엔트리: {total}건")

if __name__ == '__main__':
    main()
