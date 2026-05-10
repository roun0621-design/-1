#!/usr/bin/env python3
"""
종합기록지 생성 스크립트
사용: python3 scripts/generate_comprehensive.py <template_path> <data_json_path> <output_path>

template_path: public/template_men.xlsx 또는 public/template_women.xlsx
data_json_path: 서버에서 생성한 JSON 데이터 파일
output_path: 결과물 xlsx 경로
"""
import sys
import json
import openpyxl

def is_merged(ws, row, col):
    """Check if cell (row, col) is part of a merged range but NOT the top-left cell."""
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            if row == merged_range.min_row and col == merged_range.min_col:
                return False  # top-left: writable
            return True  # any other cell in merge: read-only
    return False  # not merged at all: writable

def safe_set(ws, row, col, value):
    """Set cell value only if it's writable (not a non-top-left merged cell)."""
    if not is_merged(ws, row, col):
        ws.cell(row, col).value = value

def main():
    if len(sys.argv) < 4:
        print("Usage: python3 generate_comprehensive.py <template> <data_json> <output>", file=sys.stderr)
        sys.exit(1)

    template_path = sys.argv[1]
    data_json_path = sys.argv[2]
    output_path = sys.argv[3]

    with open(data_json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    wb = openpyxl.load_workbook(template_path)
    ws = wb['Sheet1']

    # ---- Fill header ----
    comp = data.get('competition', {})
    safe_set(ws, 2, 2, comp.get('title', ''))
    safe_set(ws, 3, 8, comp.get('date_range', ''))
    safe_set(ws, 3, 24, comp.get('chief_judge', ''))

    # ---- Column mapping for 8 places ----
    # Position i (0-based): name_col, rec_col
    # Row offsets: 0=name+record, 1=team+wind, 2=wa_score
    PLACE_COLS = [
        (3, 5),   # 1st: C, E
        (6, 8),   # 2nd: F, H
        (9, 11),  # 3rd: I, K
        (12, 14), # 4th: L, N
        (15, 17), # 5th: O, Q
        (18, 20), # 6th: R, T
        (21, 23), # 7th: U, W
        (24, 26), # 8th: X, Z
    ]

    # ---- Event name -> template row mapping ----
    event_rows = {}
    for r in range(6, 80, 3):
        ev_name = ws.cell(r, 2).value
        if ev_name is None:
            continue
        name_str = str(ev_name).strip().replace('\n', ' ')
        if name_str.startswith('※'):
            continue
        event_rows[name_str] = r

    # ---- Fill each event ----
    events = data.get('events', [])
    for evt in events:
        template_name = evt.get('template_name', '')
        row = event_rows.get(template_name)
        if row is None:
            # Try fuzzy match (remove spaces, newlines)
            for key in event_rows:
                k_clean = key.replace(' ', '').replace('\n', '')
                t_clean = template_name.replace(' ', '').replace('\n', '')
                if k_clean == t_clean:
                    row = event_rows[key]
                    break
        if row is None:
            continue

        rankings = evt.get('rankings', [])
        is_relay = evt.get('is_relay', False)

        for i, rank_data in enumerate(rankings[:8]):
            name_col, rec_col = PLACE_COLS[i]

            # Row 0: name + record
            if is_relay:
                members = rank_data.get('members', [])
                if len(members) >= 2:
                    name_str = members[0] + ' ' + members[1]
                elif len(members) == 1:
                    name_str = members[0]
                else:
                    name_str = rank_data.get('name', '')
                # Members 3-4 go in name_col+1 on same row (D column for 1st place)
                if len(members) >= 3:
                    remaining = ' '.join(members[2:4])
                    safe_set(ws, row, name_col + 1, remaining.strip() if remaining.strip() else None)
            else:
                name_str = rank_data.get('name', '')

            record_val = rank_data.get('record', '')

            safe_set(ws, row, name_col, name_str if name_str else None)
            safe_set(ws, row, rec_col, record_val if record_val else None)

            # Row 1: team + wind
            team = rank_data.get('team', '')
            wind = rank_data.get('wind', None)

            safe_set(ws, row + 1, name_col, team if team else None)
            if wind is not None and wind != '':
                safe_set(ws, row + 1, rec_col, str(wind))
            else:
                safe_set(ws, row + 1, rec_col, None)

            # Row 2: WA score (optional)
            wa_score = rank_data.get('wa_score', None)
            if wa_score is not None and wa_score != 0 and wa_score != '':
                safe_set(ws, row + 2, rec_col, wa_score)
            else:
                safe_set(ws, row + 2, rec_col, None)

    wb.save(output_path)
    print(json.dumps({"success": True, "output": output_path}))

if __name__ == '__main__':
    main()
